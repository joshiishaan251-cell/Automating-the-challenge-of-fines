import os
import re
import shutil
import logging
import yaml
import json
import time
import subprocess
import sys
import pdfplumber
import pytesseract
import openpyxl
from openpyxl.styles import Font, Alignment
from pdf2image import convert_from_path
from datetime import datetime, timedelta
from abc import ABC, abstractmethod

# --- LOGGING SETTINGS ---
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger(__name__)

# --- CONFIGURATION CLASS ---
class Config:
    def __init__(self, config_path):
        if not os.path.exists(config_path):
            self._create_default(config_path)
            if not os.path.exists(config_path):
                logger.critical(f"Configuration file {config_path} not found and cannot be created.")
                sys.exit(1)
            
        with open(config_path, 'r', encoding='utf-8') as f:
            self.data = yaml.safe_load(f)

        # Main paths
        self.source_root = self.data.get('source_root', './downloads')
        self.archive_paths = self.data.get('archive_paths', [])
        
        # Tools
        self.poppler_path = self.data.get('poppler_path', r'C:\poppler\Library\bin\pdftotext.exe')
        self.tesseract_path = self.data.get('tesseract_path', r'C:\Program Files\Tesseract-OCR\tesseract.exe')
        
        # Mode
        self.dry_run = self.data.get('dry_run', True)
        self.verbose = self.data.get('verbose', True)
        
        # Cache
        self.cache_file = self.data.get('cache_file', 'archive_cache.json')
        self.cache_ttl = self.data.get('cache_ttl_hours', 24)

        # Name settings
        self.target_subfolder = self.data.get('target_subfolder', "3 Application")
        self.anchor_prefix = self.data.get('anchor_prefix', "! Collected")
        self.anchor_suffix = self.data.get('anchor_suffix', "_stamp")

    def _create_default(self, path):
        default_config = {
            "source_root": "./downloads",
            "archive_paths": [],
            "poppler_path": r"C:\poppler\Library\bin\pdftotext.exe",
            "tesseract_path": r"C:\Program Files\Tesseract-OCR\tesseract.exe",
            "dry_run": True,
            "verbose": True,
            "cache_file": "archive_cache.json",
            "cache_ttl_hours": 24,
            "target_subfolder": "3 Application",
            "anchor_prefix": "! Collected",
            "anchor_suffix": "_stamp"
        }
        try:
            with open(path, 'w', encoding='utf-8') as f:
                yaml.dump(default_config, f, allow_unicode=True, default_flow_style=False)
            logger.info(f"Created basic config: {path}")
        except Exception as e:
            logger.error(f"Failed to create basic config: {e}")

# --- ARCHIVE INDEXER ---
class Indexer:
    def __init__(self, config):
        self.config = config
        self.case_index = {} 
        self.uin_index = {} 
        self.timestamp = 0
        self.load_or_build()

    def load_or_build(self):
        if self._load_cache():
            logger.info("Index loaded from cache.")
        else:
            logger.info("Scanning archive...")
            self.scan_archives()
            self._save_cache()

    def _load_cache(self):
        if not os.path.exists(self.config.cache_file): return False
        try:
            with open(self.config.cache_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if time.time() - data.get('timestamp', 0) > self.config.cache_ttl * 3600:
                    return False
                self.case_index = data.get('case_index', {})
                self.uin_index = data.get('uin_index', {})
                return True
        except Exception as e:
            logger.warning(f"Cache load error: {e}")
            return False

    def _save_cache(self):
        data = {
            'timestamp': time.time(),
            'case_index': self.case_index,
            'uin_index': self.uin_index
        }
        with open(self.config.cache_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def scan_archives(self):
        count = 0
        for root_path in self.config.archive_paths:
            if not os.path.exists(root_path): continue
            
            for root, dirs, files in os.walk(root_path):
                if self.config.verbose: print(f"Index: {root[:60]}...", end='\r')
                
                for dir_name in dirs:
                    full_path = os.path.join(root, dir_name)
                    self._parse_folder_name(dir_name, full_path)
                    count += 1
                
                for file_name in files:
                    if file_name.endswith('.pdf'):
                        self._parse_file_name(file_name, root)

        if not self.config.verbose: print(f"Objects scanned: {count}")
        else: print(f"\nCompleted. Objects: {count}")

    def _parse_folder_name(self, name, path):
        case_match = re.search(r'(A|A)40[\s\-_]+(\d+)[\s\-_]+(20\d{2}|2\d)', name, re.IGNORECASE)
        if case_match:
            year = case_match.group(3)
            if len(year) == 2: year = "20" + year
            clean_case = f"{case_match.group(2)}-{year}"
            self.case_index[clean_case] = path

    def _parse_file_name(self, name, path):
        uin_match = re.search(r'(106\d{17,29}|188\d{17,29}|322\d{17,29})', name)
        if uin_match: self.uin_index[uin_match.group(1)] = path

    def find_path(self, key, key_type="case"):
        if key_type == "case": return self.case_index.get(key)
        elif key_type == "uin": return self.uin_index.get(key)
        return None

# --- BASE PROCESSOR ---
class DocumentProcessor(ABC):
    def __init__(self, indexer, config):
        self.indexer = indexer
        self.config = config

    def extract_text(self, pdf_path):
        text = ""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    extracted = page.extract_text()
                    if extracted: text += extracted + "\n"
        except Exception as e:
            logger.warning(f"pdfplumber could not read {pdf_path}: {e}")

        if self._is_text_valid(text): return self._clean_string(text)

        try:
            if os.path.exists(self.config.poppler_path):
                result = subprocess.run(
                    [self.config.poppler_path, "-layout", pdf_path, "-"],
                    capture_output=True, text=True, encoding='utf-8', errors='ignore'
                )
                if result.stdout: text = result.stdout
        except Exception as e:
            logger.warning(f"poppler could not extract text from {pdf_path}: {e}")

        if self._is_text_valid(text): return self._clean_string(text)

        try:
            if os.path.exists(self.config.tesseract_path):
                pytesseract.pytesseract.tesseract_cmd = self.config.tesseract_path
                poppler_bin = os.path.dirname(self.config.poppler_path)
                images = convert_from_path(pdf_path, first_page=1, last_page=1, poppler_path=poppler_bin)
                if images:
                    text = pytesseract.image_to_string(images[0], lang='rus+eng')
        except Exception as e:
            logger.warning(f"Tesseract OCR error for {pdf_path}: {e}")

        return self._clean_string(text)

    def _is_text_valid(self, text):
        if not text: return False
        return len(text) > 50 and bool(re.search(r'\d+', text))

    def _clean_string(self, text):
        return text.replace('\n', ' ').replace('\r', '').replace('\t', ' ')

    def move_file(self, src, dest_folder, new_name):
        if not os.path.exists(dest_folder): return False, "Target folder not found"
        
        dest_path = os.path.join(dest_folder, new_name)
        base, ext = os.path.splitext(new_name)
        counter = 1
        while os.path.exists(dest_path):
            dest_path = os.path.join(dest_folder, f"{base}_{counter}{ext}")
            counter += 1
        
        final_name = os.path.basename(dest_path)
        
        if not self.config.dry_run:
            try:
                shutil.move(src, dest_path)
                return True, final_name
            except Exception as e:
                return False, str(e)
        else:
            return True, final_name 

    def find_subfolder(self, root, sub_name):
        try:
            for d in os.listdir(root):
                if os.path.isdir(os.path.join(root, d)):
                    if sub_name.lower() in d.lower():
                        return os.path.join(root, d)
        except Exception as e:
            logger.warning(f"Error searching subfolder '{sub_name}' in {root}: {e}")
        return None

    @abstractmethod
    def process(self, pdf_path, text):
        pass

# --- ACT PROCESSOR ---
class ActProcessor(DocumentProcessor):
    def process(self, pdf_path, text):
        found_uins = []
        potential_uins = re.findall(r'(106[\d\s\-]{17,35}|188[\d\s\-]{17,35})', text)
        for p_uin in potential_uins:
            digits = re.sub(r'\D', '', p_uin)
            if 20 <= len(digits) <= 29: found_uins.append(digits)

        found_cases = []
        # Support both Russian and English keywords
        keyword_matches = re.finditer(r'(?:Case|N|№|Decision|Решение|Постановление|Определение)', text, re.IGNORECASE)
        for match in keyword_matches:
            snippet = text[match.end():match.end()+50]
            # Fixed O/0, Z/3 and other common OCR errors
            table = str.maketrans("ЗОБЧАSsg", "30644559") 
            snippet_fixed = snippet.translate(table)
            matches = re.findall(r'(?<!\d)(\d{5,7})\D{0,5}(202\d|2\d)(?!\d)', snippet_fixed)
            for m in matches:
                found_cases.append(f"{m[0]}-{self._normalize_year(m[1])}")

        std_matches = re.findall(r'40[\s\-\/\\]+(\d{5,7})[\s\-\/\\]+(\d{2,4})', text)
        for m in std_matches:
            found_cases.append(f"{m[0]}-{self._normalize_year(m[1])}")

        found_cases = list(set(found_cases))
        found_uins = list(set(found_uins))

        target_path = None
        match_key = "---"
        match_type = "---"

        for uin in found_uins:
            path = self.indexer.find_path(uin, "uin")
            if path:
                target_path, match_key, match_type = path, uin, "UIN"
                break
        
        if not target_path:
            for case in found_cases:
                path = self.indexer.find_path(case, "case")
                if path:
                    target_path, match_key, match_type = path, case, "Case"
                    break

        if not target_path:
            all_keys = found_cases + found_uins
            return {"status": "Not Found (Archive)", "match_key": all_keys if all_keys else "---"}

        final_target_dir = target_path
        sub_dir = self.find_subfolder(target_path, self.config.target_subfolder)
        if sub_dir: final_target_dir = sub_dir
        
        new_name = None
        try:
            files_in_folder = os.listdir(final_target_dir) if os.path.exists(final_target_dir) else []
            for file in files_in_folder:
                if file.startswith(self.config.anchor_prefix) and file.endswith(".pdf"):
                    # Support both Russian and English A40
                    if "A40" in file.upper() or "А40" in file.upper(): 
                        if self.config.anchor_suffix not in file:
                            anchor_name = os.path.splitext(file)[0]
                            new_name = f"{anchor_name}{self.config.anchor_suffix}.pdf"
                            break
        except Exception as e:
            logger.warning(f"Error searching anchor file in {final_target_dir}: {e}")

        if not new_name:
            current_date = datetime.now().strftime("%Y%m%d")
            case_part = f"A40-{match_key}" if "A40" not in str(match_key).upper() else match_key
            if match_type == "UIN" and found_cases:
                case_part = f"A40-{found_cases[0]}"
            new_name = f"{self.config.anchor_prefix} {case_part}_{current_date}{self.config.anchor_suffix}.pdf"

        success, final_name = self.move_file(pdf_path, final_target_dir, new_name)
        status = "Success" if success else f"Error: {final_name}"
        return {"status": status, "match_type": match_type, "match_key": match_key, 
                "target_path": final_target_dir, "new_name": final_name}
    
    def _normalize_year(self, year_str):
        return "20" + year_str if len(year_str) == 2 else year_str

# --- CHECK PROCESSOR ---
class CheckProcessor(DocumentProcessor):
    def process(self, pdf_path, text):
        uin = None
        date_str = None
        is_payment_order = "PAYMENT ORDER" in text.upper() or "ПЛАТЕЖНОЕ ПОРУЧЕНИЕ" in text.upper()
        candidates = []

        if is_payment_order:
            # Context matches for "resolution", "penalty", etc.
            context_matches = re.findall(r'(?:пост|постановл|штраф|provision|resolution|penalty)[а-я\.]*[:\s№]*(\d[\d\s]{15,35})', text, re.IGNORECASE)
            for m in context_matches: candidates.append(re.sub(r'\D', '', m))
            uin_matches = re.findall(r'УИН|UIN[:\s№]*(\d[\d\s]{15,35})', text, re.IGNORECASE)
            for m in uin_matches: candidates.append(re.sub(r'\D', '', m))
            raw_long_digits = re.findall(r'(?<!\d)(\d[\d\s\-]{18,35}\d)(?!\d)', text)
            for m in raw_long_digits: candidates.append(re.sub(r'\D', '', m))
        else:
            m = re.search(r'(?:УИН|Идентификатор|Постановление|UIN|Identifier|Resolution)[:\.\s]*(\d{20,29})', text, re.IGNORECASE)
            if m: candidates.append(m.group(1))

        valid_candidates = []
        for digits in candidates:
            if not (20 <= len(digits) <= 25): continue
            # Skip common bank accounts
            if digits.startswith(("408", "407", "406", "405", "301", "302", "032", "031", "401")): continue
            valid_candidates.append(digits)
        
        if valid_candidates:
            valid_candidates = list(set(valid_candidates))
            prio = next((x for x in valid_candidates if x.startswith(("106", "188", "322"))), None)
            uin = prio if prio else valid_candidates[0]

        m_date = re.search(r'(\d{2}[\.\/]\d{2}[\.\/]\d{4})', text)
        if m_date: date_str = m_date.group(1).replace('/', '.')

        if not uin: return {"status": "Not Found (No UIN)", "match_key": "---"}

        target_path = self.indexer.find_path(uin, "uin")
        if not target_path: return {"status": "Not Found (Archive)", "match_key": uin}

        final_target_dir = target_path
        sub_dir = self.find_subfolder(target_path, self.config.target_subfolder)
        if sub_dir: final_target_dir = sub_dir
            
        safe_date = date_str.replace('.', '-') if date_str else "no_date"
        ext = os.path.splitext(pdf_path)[1]
        new_name = f"Check_{safe_date}_{uin}{ext}"
        
        success, final_name = self.move_file(pdf_path, final_target_dir, new_name)
        status = "Success" if success else f"Error: {final_name}"
        return {"status": status, "match_type": "UIN", "match_key": uin, 
                "target_path": final_target_dir, "new_name": final_name}

# --- MAIN CONTROLLER ---
def main():
    print("--- LAUNCHING UNIVERSAL SORTER ---")
    config = Config('config.yaml')
    indexer = Indexer(config)
    
    act_proc = ActProcessor(indexer, config)
    check_proc = CheckProcessor(indexer, config)
    
    if config.dry_run:
        print("[!] DRY RUN MODE (No moving)")

    files_to_process = []
    print(f"Scanning incoming folder: {config.source_root}")
    for root, dirs, files in os.walk(config.source_root):
        for filename in files:
            if filename.lower().endswith('.pdf'):
                files_to_process.append(os.path.join(root, filename))

    total_files = len(files_to_process)
    print(f"Files found for processing: {total_files}")
    
    report_data = []

    if config.verbose:
        print("-" * 80)
        print(f"{'File':<20} | {'Type':<10} | {'Status':<15} | {'Key':<20}")
        print("-" * 80)

    for i, file_path in enumerate(files_to_process, 1):
        filename = os.path.basename(file_path)
        if not config.verbose: print(f"Processing: {i}/{total_files} - {filename[:30]}...", end='\r')

        text = act_proc.extract_text(file_path)
        if not text:
            if config.verbose: print(f"{filename[:20]:<20} | {'ERR':<10} | Empty Text")
            report_data.append([filename, "Unknown", "Empty Text", "", "", ""])
            continue

        is_act = re.search(r'(Решение|Постановление|Приказ|Определение|Decision|Resolution|Order)', text, re.IGNORECASE)
        is_payment = re.search(r'(Чек|Платежное поручение|Сбербанк|Банк|Оплата|Check|Payment|Receipt|Bank)', text, re.IGNORECASE)
        
        processor = None
        p_type = "Unknown"
        if is_act and not is_payment:
            processor, p_type = act_proc, "Act"
        elif is_payment:
            processor, p_type = check_proc, "Check"
        elif is_act: 
            processor, p_type = act_proc, "Act"
            
        if not processor:
            # ZERO TRUST: Unknown documents must be ignored, not treated as fallbacks.
            if config.verbose: print(f"{filename[:20]:<20} | {'Unknown':<10} | Ignored | ---")
            report_data.append([filename, "Unknown", "Ignored (Zero Trust)", "---", "", ""])
            continue

        res = processor.process(file_path, text)
        
        raw_key = res.get('match_key', '---')
        key_str = str(raw_key).replace("['", "").replace("']", "")
        
        if config.verbose:
            print(f"{filename[:20]:<20} | {p_type:<10} | {res.get('status'):<15} | {key_str[:20]}")
        
        report_data.append([
            filename, p_type, res.get('status'), key_str, 
            res.get('new_name', ''), res.get('target_path', '')
        ])

    print(f"\nProcessing complete. Total: {total_files}")
    
    # --- SAVING TO REAL EXCEL (.xlsx) ---
    current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    xlsx_filename = f'universal_report_{current_time}.xlsx'
    # Use absolute path relative to the script's execution so it's predictable
    xlsx_file = os.path.abspath(os.path.join(os.getcwd(), xlsx_filename))
    
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sorting Report"
        
        # 1. Headers
        headers = ["File", "Type", "Status", "Key", "New Name", "Path"]
        ws.append(headers)
        
        # 2. Width and style settings (350 pixels ~ 50 chars)
        column_widths = [50, 15, 25, 35, 50, 80]
        
        # Styles
        header_font = Font(bold=True)
        green_font = Font(color="006100")
        red_font = Font(color="9C0006")
        
        for i, width in enumerate(column_widths, 1):
            col_letter = openpyxl.utils.get_column_letter(i)
            ws.column_dimensions[col_letter].width = width
            cell = ws.cell(row=1, column=i)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # 3. Writing data
        for row in report_data:
            # Cleaning data
            clean_row = []
            for cell in row:
                s_cell = str(cell)
                # Removing CSV hacks if present
                if s_cell.startswith('="') and s_cell.endswith('"'):
                    clean_row.append(s_cell.replace('="', '').replace('"', ''))
                else:
                    clean_row.append(cell)
            
            ws.append(clean_row)
            
            # Color statuses
            last_row = ws.max_row
            status_cell = ws.cell(row=last_row, column=3)
            
            if "Success" in str(status_cell.value):
                for col in range(1, 7):
                    ws.cell(row=last_row, column=col).font = green_font
            elif "Not Found" in str(status_cell.value):
                status_cell.font = red_font

        wb.save(xlsx_file)
        print(f"Excel report saved: {xlsx_file}")
        
    except Exception as e:
        print(f"Excel save error: {e}")

if __name__ == "__main__":
    main()