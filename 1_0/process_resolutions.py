import os
import re
import json
import yaml
import sys
import tkinter as tk
from tkinter import messagebox, simpledialog
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, Any, List

import fitz  # PyMuPDF
from PIL import Image, ImageTk, ImageEnhance
import pytesseract
from pdf2image import convert_from_path

# Add project root to sys.path to access INDEX_UIN
PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.append(str(PROJECT_ROOT))

try:
    from INDEX_UIN.db_manager import DBManager
except ImportError:
    DBManager = None

# --- CONFIGURATION LOADING ---
def load_config():
    config_path = Path(__file__).resolve().parent / "resolution_config.yaml"
    with open(config_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)

CONFIG = load_config()

# --- PATTERNS ---
RE_UIN = re.compile(r"\b(\d{20,25})\b")
# Folder pattern: Letter (Cyrillic or Latin) + 1-3 digits (e.g. A1, B13, V123)
FOLDER_RE = re.compile(CONFIG.get('folder_pattern', r'^[А-ЯA-Z]\d{1,3}.*'))

class ZoomableImage(tk.Frame):
    """Component to display PDF page with zoom and pan."""
    def __init__(self, master, pil_image, **kwargs):
        super().__init__(master, **kwargs)
        self.pil_image = pil_image
        self.scale = 1.0
        self.delta = 1.3
        self.vbar = tk.Scrollbar(self, orient='vertical')
        self.hbar = tk.Scrollbar(self, orient='horizontal')
        self.vbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.hbar.pack(side=tk.BOTTOM, fill=tk.X)
        self.canvas = tk.Canvas(self, bg='gray', xscrollcommand=self.hbar.set, yscrollcommand=self.vbar.set)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.vbar.config(command=self.canvas.yview)
        self.hbar.config(command=self.canvas.xview)
        self.canvas.bind('<ButtonPress-1>', self.move_start)
        self.canvas.bind('<B1-Motion>', self.move_move)
        self.canvas.bind('<MouseWheel>', self.wheel)
        self.show_image()

    def show_image(self):
        if self.pil_image is None: return
        w, h = self.pil_image.size
        new_size = int(w * self.scale), int(h * self.scale)
        imagetk = ImageTk.PhotoImage(self.pil_image.resize(new_size, Image.Resampling.LANCZOS))
        self.canvas.delete("all")
        self.imageid = self.canvas.create_image(0, 0, anchor='nw', image=imagetk)
        self.canvas.imagetk = imagetk
        self.canvas.config(scrollregion=self.canvas.bbox('all'))

    def move_start(self, event):
        self.canvas.scan_mark(event.x, event.y)

    def move_move(self, event):
        self.canvas.scan_dragto(event.x, event.y, gain=1)

    def wheel(self, event):
        if event.delta > 0: self.scale *= self.delta
        elif event.delta < 0: self.scale /= self.delta
        self.scale = max(0.1, min(self.scale, 5.0))
        self.show_image()

class DuplicateDialog(tk.Toplevel):
    """Dialog to show duplicate paths and offer deletion."""
    def __init__(self, master, uin, paths):
        super().__init__(master)
        self.uin = uin
        self.paths = paths
        self.action = None # 'delete' or 'skip'
        
        self.title(f"Duplicate Found: {uin}")
        self.geometry("600x400")
        self.attributes('-topmost', True)
        
        tk.Label(self, text=f"UIN {uin} is already indexed!", font=("Arial", 12, "bold"), fg="red", pady=10).pack()
        tk.Label(self, text="This number appears in the following files:").pack()
        
        frame = tk.Frame(self)
        frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        text = tk.Text(frame, height=10, font=("Consolas", 9))
        text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar = tk.Scrollbar(frame, command=text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        text.config(yscrollcommand=scrollbar.set)
        
        for p in paths:
            text.insert(tk.END, f"{p}\n")
        text.config(state=tk.DISABLED)

        btn_copy = tk.Button(self, text="Copy Paths", command=lambda: self.copy_paths(paths))
        btn_copy.pack(pady=5)

        tk.Label(self, text="Delete current file?", pady=10).pack()
        
        btn_fm = tk.Frame(self)
        btn_fm.pack(pady=10)
        
        tk.Button(btn_fm, text="Delete to Trash and CONTINUE", bg="#f44336", fg="white", 
                  padx=10, command=self.on_delete).pack(side=tk.LEFT, padx=10)
        tk.Button(btn_fm, text="Skip Folder (STOP)", bg="#9E9E9E", fg="white",
                  padx=10, command=self.on_stop).pack(side=tk.LEFT, padx=10)

    def copy_paths(self, paths):
        self.clipboard_clear()
        self.clipboard_append("\n".join(paths))
        messagebox.showinfo("Info", "Paths copied to clipboard")

    def on_delete(self):
        self.action = 'delete'
        self.destroy()

    def on_stop(self):
        self.action = 'stop'
        self.destroy()

class ResolutionForm(tk.Toplevel):
    """GUI Window for a single resolution processing."""
    def __init__(self, master, pdf_path, initial_uin, initial_date, initial_veh, is_duplicate):
        super().__init__(master)
        self.pdf_path = pdf_path
        self.initial_uin = initial_uin
        self.initial_date = initial_date
        self.initial_veh = initial_veh
        self.is_duplicate = is_duplicate
        self.result = None
        
        self.title(f"Processing: {os.path.basename(pdf_path)}")
        self.state('zoomed')
        
        self._setup_ui()
        if is_duplicate:
            messagebox.showerror("Duplicate", f"UIN {initial_uin} is already in the main index!")
            self.destroy()

    def _setup_ui(self):
        # Left side: Viewer
        self.frame_viewer = tk.Frame(self, bg="#333")
        self.frame_viewer.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        try:
            doc = fitz.open(self.pdf_path)
            pix = doc[0].get_pixmap(matrix=fitz.Matrix(2, 2))
            pil_image = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            self.viewer = ZoomableImage(self.frame_viewer, pil_image)
            self.viewer.pack(fill=tk.BOTH, expand=True)
            doc.close()
        except Exception as e:
            tk.Label(self.frame_viewer, text=f"PDF Load Error: {e}", fg="white").pack()

        # Right side: Form
        self.frame_form = tk.Frame(self, width=400, padx=20, pady=20)
        self.frame_form.pack(side=tk.RIGHT, fill=tk.Y)
        self.frame_form.pack_propagate(False)

        tk.Label(self.frame_form, text="Resolution Verification", font=("Arial", 14, "bold")).pack(pady=(0, 20))

        tk.Label(self.frame_form, text="Resolution Number (UIN):").pack(anchor="w")
        self.ent_uin = tk.Entry(self.frame_form, font=("Arial", 12))
        self.ent_uin.pack(fill=tk.X, pady=(0, 15))
        if self.initial_uin:
            self.ent_uin.insert(0, self.initial_uin)

        tk.Label(self.frame_form, text="Resolution Date:").pack(anchor="w")
        self.ent_date = tk.Entry(self.frame_form, font=("Arial", 12))
        self.ent_date.pack(fill=tk.X, pady=(0, 15))
        if self.initial_date:
            self.ent_date.insert(0, self.initial_date)

        tk.Label(self.frame_form, text="Vehicle Plate Number:").pack(anchor="w")
        self.ent_veh = tk.Entry(self.frame_form, font=("Arial", 12))
        self.ent_veh.pack(fill=tk.X, pady=(0, 20))
        if self.initial_veh:
            self.ent_veh.insert(0, self.initial_veh)
        self.ent_veh.focus_set()

        btn_save = tk.Button(self.frame_form, text="SAVE", bg="#4CAF50", fg="white", 
                             font=("Arial", 12, "bold"), command=self.on_save)
        btn_save.pack(fill=tk.X, pady=10)

        btn_skip = tk.Button(self.frame_form, text="SKIP", bg="#9E9E9E", fg="white", 
                              command=self.on_skip)
        btn_skip.pack(fill=tk.X)

        self.bind("<Return>", lambda e: self.on_save())

    def on_save(self):
        uin = self.ent_uin.get().strip()
        date_str = self.ent_date.get().strip()
        veh = self.ent_veh.get().strip()
        if not uin:
            messagebox.showwarning("Warning", "UIN cannot be empty")
            return
        self.result = {"uin": uin, "date": date_str, "veh": veh}
        self.destroy()

    def on_skip(self):
        self.result = None
        self.destroy()

class ResolutionProcessor:
    def __init__(self):
        self.config = CONFIG
        self.root = tk.Tk()
        self.root.withdraw()
        
        db_path = self.config.get('central_db_path')
        self.db_manager = DBManager(db_path) if DBManager and db_path else None
        
        if os.path.exists(self.config.get('tesseract_cmd', '')):
            pytesseract.pytesseract.tesseract_cmd = self.config['tesseract_cmd']

    def is_in_central_db(self, uin: str) -> bool:
        if not self.db_manager:
            return False
        with self.db_manager._connection() as conn:
            row = conn.execute("SELECT 1 FROM uins WHERE number = ?", (uin,)).fetchone()
            return row is not None

    def get_duplicate_paths(self, uin: str) -> List[str]:
        """Fetch all file paths where this UIN is found from central index."""
        if not self.db_manager:
            return []
        paths = []
        with self.db_manager._connection() as conn:
            query = """
                SELECT a.path, o.filename 
                FROM occurrences o
                JOIN uins u ON o.uin_id = u.id
                JOIN archives a ON o.archive_id = a.id
                WHERE u.number = ?
            """
            rows = conn.execute(query, (uin,)).fetchall()
            for r in rows:
                paths.append(os.path.join(r[0], r[1]))
        return paths

    def move_to_trash(self, file_path: Path):
        """Move file to Windows Recycle Bin using PowerShell."""
        import subprocess
        try:
            ps_safe_path = str(file_path).replace("'", "''")
            cmd = f'powershell.exe -Command "Add-Type -AssemblyName Microsoft.VisualBasic; [Microsoft.VisualBasic.FileIO.FileSystem]::DeleteFile(\'{ps_safe_path}\', \'OnlyErrorDialogs\', \'SendToRecycleBin\')"'
            subprocess.run(cmd, check=True, shell=True)
        except Exception as e:
            print(f"Error moving to trash (PS): {e}")
            try:
                file_path.unlink()
            except Exception as e2:
                print(f"Error deleting file: {e2}")
                return False

        # Verify file was actually removed from disk
        if file_path.exists():
            print(f"  WARNING: File {file_path.name} still exists after deletion operation!")
            return False  # Signal failure even if PS returned OK

        print(f"File moved to trash: {file_path.name}")
        return True

    def update_existing_report(self, uin: str, archive_path: str, filename: str):
        """Append newly indexed UIN to the latest Excel report and highlight it in red."""
        import glob
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font
        except ImportError:
            print("openpyxl not installed, report update skipped.")
            return

        reports_dir = PROJECT_ROOT / "reports"
        pattern = str(reports_dir / "UIN_Duplicates_Report_*.xlsx")
        reports = glob.glob(pattern)
        
        if not reports:
            print("Report to update not found (UIN_Duplicates_Report_*.xlsx files missing).")
            return
            
        latest_report = max(reports, key=os.path.getmtime)
        print(f"Updating report: {os.path.basename(latest_report)}")
        
        try:
            wb = openpyxl.load_workbook(latest_report)
            if "All UINs" in wb.sheetnames or "Все УИН" in wb.sheetnames:
                sheet_name = "All UINs" if "All UINs" in wb.sheetnames else "Все УИН"
                ws = wb[sheet_name]
                
                now_str = datetime.now().isoformat()
                row_data = [uin, filename, archive_path, now_str, ""]
                ws.append(row_data)
                
                # Highlight the newly appended row
                red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                white_font = Font(color='FFFFFFFF', bold=True)
                new_row_idx = ws.max_row
                
                for cell in ws[new_row_idx]:
                    cell.fill = red_fill
                    cell.font = white_font
                    
                # Use tempfile and atomic replace to avoid corrupted .xlsx if interrupted
                import tempfile
                fd, tmp_path = tempfile.mkstemp(dir=os.path.dirname(latest_report), suffix=".tmp.xlsx")
                os.close(fd)
                
                wb.save(tmp_path)
                os.replace(tmp_path, latest_report)
                print(f"Row added to report and highlighted in red.")
            else:
                print("'All UINs' sheet not found in the report.")
        except PermissionError:
            messagebox.showwarning("Error", "Excel report file is open! Please close it. UIN recording skipped.")
            if 'tmp_path' in locals() and os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception as e:
            print(f"Error updating Excel report: {e}")

    def perform_ocr_on_pdf(self, pdf_path: Path) -> str:
        """Convert PDF to images and perform OCR on each page."""
        try:
            from pdf2image import convert_from_path
            import pytesseract
            
            print(f"Performing OCR for: {pdf_path.name}...")
            # Use poppler path from config if available. Limit OCR to first 2 pages.
            poppler_path = self.config.get('poppler_bin')
            images = convert_from_path(str(pdf_path), poppler_path=poppler_path, last_page=2)
            
            full_text = ""
            for i, img in enumerate(images):
                # Enhance image for better OCR
                # img = img.convert('L') # Greyscale
                text = pytesseract.image_to_string(img, lang='rus+eng')
                full_text += text + "\n"
            
            return full_text
        except Exception as e:
            print(f"OCR Error: {e}")
            return ""

    def extract_doc_info(self, pdf_path: Path) -> Dict[str, str]:
        """Extract UIN, Date, and Vehicle from filename or PDF text layer/OCR."""
        info = {
            "uin": "",
            "date": datetime.now().strftime("%d.%m.%Y"),
            "veh": ""
        }
        
        # 1. UIN From filename (if it's already a UIN)
        match_uin = RE_UIN.search(pdf_path.name)
        if match_uin:
            info["uin"] = match_uin.group(1)
        
        # 2. Try text layer first
        text = ""
        try:
            doc = fitz.open(pdf_path)
            text = "".join(page.get_text() for page in doc)
            doc.close()
        except:
            pass

        # 3. If text layer is empty or missing data, try OCR if enabled
        if self.config.get('perform_ocr', True):
            # Check against skip list from config (e.g. 'Delivered' files are not resolutions)
            skip_patterns = self.config.get('ocr_skip_patterns', [])
            fname_lower = pdf_path.name.lower()
            should_skip_ocr = any(fname_lower.startswith(p.lower()) for p in skip_patterns)

            if should_skip_ocr:
                print(f"  OCR skipped (by config) for: {pdf_path.name}")
            elif len(text.strip()) < 100 or not info["uin"]:
                # Heuristic: typical resolution text layer has at least 100 chars
                ocr_text = self.perform_ocr_on_pdf(pdf_path)
                text += "\n" + ocr_text

        # 4. Extract from accumulated text
        if text:
            if not info["uin"]:
                match = RE_UIN.search(text)
                if match:
                    info["uin"] = match.group(1)
            
            # Simple Date Extraction DD.MM.YYYY
            # OCR might add spaces between digits: "1 2 . 0 3 . 2 0 2 6", so we clean first
            date_clean_text = text.replace(" ", "")
            date_match = re.search(r'(\d{2}\.\d{2}\.\d{4})', date_clean_text)
            if date_match:
                info["date"] = date_match.group(1)
                
            # Simple Vehicle Plate Extraction (A123BC77)
            clean_text = text.replace(" ", "").replace("\n", "")
            # Russian plates + English homoglyphs
            rus_eng_chars = r'[АВЕКМНОРСТУХABEKMHOPCTYX]'
            veh_match = re.search(f'({rus_eng_chars}\\d{{3}}{rus_eng_chars}{{2}}\\d{{2,3}})', clean_text)
            if veh_match:
                raw_plate = veh_match.group(1).upper()
                # Normalize: Latin OCR lookalikes -> Cyrillic
                latin_to_cyr = {
                    'A': 'А', 'B': 'В', 'C': 'С', 'E': 'Е', 'H': 'Н',
                    'K': 'К', 'M': 'М', 'O': 'О', 'P': 'Р', 'T': 'Т',
                    'X': 'Х', 'Y': 'У'
                }
                normalized = ''.join(latin_to_cyr.get(ch, ch) for ch in raw_plate)
                if normalized != raw_plate:
                    print(f"  Normalized vehicle plate: {raw_plate} -> {normalized}")
                info["veh"] = normalized
            
        return info

    def load_existing_database(self, folder_path: Path) -> Dict[str, Any]:
        """Load database.json if it exists, otherwise return default structure."""
        db_path = folder_path / "BASE" / "database.json"
        if db_path.exists():
            try:
                with open(db_path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception as e:
                print(f"Database Read Error: {e}")
        
        # Default structure
        now_str = datetime.now().strftime("%d.%m.%Y")
        db = {
            "CaseNumbers": [],
            "VehicleNumbers": [],
            "PochtaNumber": "",
            "DataNumber": now_str,
            "FolderName": folder_path.name,
            "NumberDeloForFileName": "", 
            "NumberDeloForDoc": "",
            "Sudilo": ""
        }
        db.update(self.config.get('static_data', {}))
        return db

    def update_database_json(self, folder_path: Path, uin: str, date_str: str, veh: str):
        """Add a single entry to database.json incrementally (Atomic Save)"""
        import tempfile
        base_dir = folder_path / "BASE"
        base_dir.mkdir(exist_ok=True)
        db_path = base_dir / "database.json"
        
        db = self.load_existing_database(folder_path)
        
        case_entry = f"№ {uin} from {date_str}"
        
        if case_entry not in db["CaseNumbers"]:
            db["CaseNumbers"].append(case_entry)
            db["VehicleNumbers"].append(veh)
            # Add or update individual VehicleNumberX
            idx = len(db["VehicleNumbers"])
            db[f"VehicleNumber{idx}"] = veh
            
        # Atomic Write
        fd, temp_path = tempfile.mkstemp(dir=str(base_dir), prefix="db_tmp_", suffix=".json")
        try:
            with os.fdopen(fd, 'w', encoding='utf-8') as f:
                json.dump(db, f, ensure_ascii=False, indent=4)
            os.replace(temp_path, db_path)
            print(f"Database successfully updated: {uin}")
        except Exception as e:
            if os.path.exists(temp_path): os.remove(temp_path)
            print(f"CRITICAL SAVE ERROR: {e}")
            raise

    def find_active_folder(self) -> Optional[Path]:
        """List ALL matching folders in scan_root in terminal and ask user to pick by number."""
        root = Path(self.config.get('scan_root', '.'))
        if not root.exists():
            print(f"Error: Scan root not found: {root}")
            return None

        candidates = sorted(
            [item for item in root.iterdir() if item.is_dir() and FOLDER_RE.match(item.name)],
            key=lambda p: p.name
        )

        if not candidates:
            print(f"In folder {root}, no folders matching the pattern (e.g., A1, B13) were found.")
            return None

        if len(candidates) == 1:
            print(f"One folder found: {candidates[0].name} — automatically selected.")
            return candidates[0]

        # Terminal selection
        print("\nAvailable folders for processing:")
        print("-" * 40)
        for i, c in enumerate(candidates, start=1):
            # Count already-processed UINs for quick status
            db_path = c / "BASE" / "database.json"
            done = 0
            if db_path.exists():
                try:
                    import json as _json
                    with open(db_path, 'r', encoding='utf-8') as f:
                        _db = _json.load(f)
                    done = len(_db.get("CaseNumbers", []))
                except Exception:
                    pass
            status = f"{done}/4 processed" if done else "new"
            print(f"  {i}. {c.name}  [{status}]")
        print("-" * 40)

        while True:
            try:
                raw = input(f"Enter folder number (1-{len(candidates)}): ").strip()
                choice = int(raw)
                if 1 <= choice <= len(candidates):
                    selected = candidates[choice - 1]
                    print(f"Selected folder: {selected.name}\n")
                    return selected
                else:
                    print(f"  Enter a number between 1 and {len(candidates)}.")
            except ValueError:
                print("  Please enter a number.")
            except (EOFError, KeyboardInterrupt):
                print("\nSelection cancelled.")
                return None


    def extract_delivery_info(self, folder_path: Path) -> Dict[str, str]:
        """Look for 'Delivered' file and extract DataNumber/PochtaNumber from its name."""
        info = {}
        try:
            files = os.listdir(folder_path)
        except Exception as e:
            print(f"Error reading folder for 'Delivered' search: {e}")
            return info

        for fname in files:
            # Check for Russian and English prefixes
            if fname.lower().startswith(('вручено', 'delivered')):
                print(f"  Delivery file found: {fname}")
                # Pattern: Delivered 20.02.2026 ... RPO 63097717173898
                date_match = re.search(r'(?:Delivered|Вручено)\s+(\d{2}\.\d{2}\.\d{4})', fname, re.IGNORECASE)
                if not date_match:
                    # Fallback: any date in the filename
                    date_match = re.search(r'(\d{2}\.\d{2}\.\d{4})', fname)

                rpo_match = re.search(r'(?:RPO|РПО)\s*(\d+)', fname, re.IGNORECASE)

                if date_match:
                    info["DataNumber"] = date_match.group(1)
                    print(f"  Extracted delivery date: {info['DataNumber']}")
                if rpo_match:
                    info["PochtaNumber"] = rpo_match.group(1)
                    print(f"  Extracted RPO number: {info['PochtaNumber']}")
                break 
        
        if not info:
            print("  'Delivered...' file not found in folder. DataNumber and PochtaNumber not updated.")
        return info

    def run(self):
        folder = self.find_active_folder()
        if not folder:
            messagebox.showinfo("Info", "Active folder not selected or not found.")
            return

        print(f"Processing folder: {folder}")
        
        # Ensure BASE folder exists
        base_dir = folder / "BASE"
        base_dir.mkdir(exist_ok=True)
        
        # Immediate Initialization of database.json if missing
        db_path = base_dir / "database.json"
        
        # Load current state
        db = self.load_existing_database(folder)
        
        # If the file didn't exist, this forces it to be created immediately
        if not db_path.exists():
            try:
                with open(db_path, 'w', encoding='utf-8') as f:
                    json.dump(db, f, ensure_ascii=False, indent=4)
                print(f"Initialized database file: {db_path}")
            except Exception as e:
                print(f"Database Initialization Error: {e}")

        processed_uins = []
        for entry in db.get("CaseNumbers", []):
            m = RE_UIN.search(entry)
            if m: processed_uins.append(m.group(1))

        # --- Phase 1: Extract Delivery Info from "Delivered" file ---
        print("\n[Phase 1] Searching for delivery file ('Delivered...')...")
        delivery_info = self.extract_delivery_info(folder)
        if delivery_info:
            updated = False
            if "DataNumber" in delivery_info:
                old_val = db.get("DataNumber", "None")
                db["DataNumber"] = delivery_info["DataNumber"]
                if old_val != delivery_info["DataNumber"]:
                    print(f"  -> DataNumber changed: {old_val} -> {delivery_info['DataNumber']}")
                    updated = True
            if "PochtaNumber" in delivery_info:
                old_val = db.get("PochtaNumber", "None")
                db["PochtaNumber"] = delivery_info["PochtaNumber"]
                if old_val != delivery_info["PochtaNumber"]:
                    print(f"  -> PochtaNumber changed: {old_val} -> {delivery_info['PochtaNumber']}")
                    updated = True

            # Always save to DB to ensure file is up to date
            try:
                with open(db_path, 'w', encoding='utf-8') as f:
                    json.dump(db, f, ensure_ascii=False, indent=4)
                if updated:
                    print("  Database updated with data from 'Delivered' file.")
                else:
                    print("  Delivery data matches database, no changes.")
            except Exception as e:
                print(f"  Error saving database after updating delivery data: {e}")

        pdf_files = sorted(list(folder.glob("*.pdf")))

        # Load skip patterns from config
        skip_patterns = self.config.get('ocr_skip_patterns', [])

        to_process = []
        for pdf in pdf_files:
            fname_lower = pdf.name.lower()
            # Completely exclude files matching skip patterns
            if any(fname_lower.startswith(p.lower()) for p in skip_patterns):
                print(f"Skipping (not a resolution, by config): {pdf.name}")
                continue
            # Check if file is already named as a UIN and present in DB
            match = RE_UIN.search(pdf.name)
            if match and match.group(1) in processed_uins:
                print(f"Skipping (already in database): {pdf.name}")
                continue
            to_process.append(pdf)

        # Limit to resolutions_count NEW resolutions
        target_count = int(self.config.get('resolutions_count', 4))
        to_process = to_process[:max(0, target_count - len(processed_uins))]

        if not to_process:
            if len(processed_uins) >= target_count:
                messagebox.showinfo("Info", f"In folder {folder.name}, {target_count} resolutions have already been processed.")
            else:
                messagebox.showinfo("Info", f"No new PDF files for processing in folder {folder.name}.")

        for pdf_path in to_process:
            doc_info = self.extract_doc_info(pdf_path)
            uin = doc_info["uin"]

            # Check duplicates in central index
            is_dupe = False
            if uin and self.is_in_central_db(uin):
                paths = self.get_duplicate_paths(uin)
                dlg = DuplicateDialog(self.root, uin, paths)
                self.root.wait_window(dlg)

                if dlg.action == 'delete':
                    if self.move_to_trash(pdf_path):
                        print(f"File {pdf_path.name} moved to trash. Continuing.")
                        continue
                    else:
                        messagebox.showerror("Error", "Failed to delete file.")
                        return
                else:
                    return

            # Show GUI for verification/input
            form = ResolutionForm(self.root, str(pdf_path), uin, doc_info["date"], doc_info["veh"], is_dupe)
            self.root.wait_window(form)

            if form.result:
                final_uin = form.result['uin']

                update_mode = self.config.get('update_index_mode', 'ask').lower()
                save_to_index = False

                if self.db_manager:
                    if update_mode == 'always':
                        save_to_index = True
                    elif update_mode == 'ask':
                        save_to_index = messagebox.askyesno(
                            "Add to Index",
                            f"Add all information for this file to the index database?\n(UIN: {final_uin})"
                        )

                if save_to_index and self.db_manager:
                    import hashlib
                    folder_hash = hashlib.md5(str(folder).encode('utf-8')).hexdigest()
                    archive_id, _ = self.db_manager.get_or_update_archive_atomic(str(folder), folder_hash)
                    self.db_manager.add_uin_occurrence(final_uin, archive_id, f"{final_uin}.pdf")
                    print(f"UIN {final_uin} added to central index.")
                    self.update_existing_report(final_uin, str(folder), f"{final_uin}.pdf")

                # Update local database
                self.update_database_json(folder, final_uin, form.result['date'], form.result['veh'])

                # Rename file
                new_name = folder / f"{final_uin}.pdf"
                try:
                    if pdf_path != new_name:
                        if new_name.exists():
                            new_name = folder / f"{form.result['uin']}_{datetime.now().strftime('%H%M%S')}.pdf"
                        pdf_path.rename(new_name)
                        print(f"Renamed: {pdf_path.name} -> {new_name.name}")
                except Exception as e:
                    print(f"Rename Error: {e}")
            else:
                print(f"Skipped by user: {pdf_path.name}")

        # --- Phase 2: Generate Statement Documents (.docx, .pdf) ---
        if messagebox.askyesno("Document Generation", "Folder processed. Generate court statement (.docx and .pdf)?"):
            try:
                import generate_docs
                generate_docs.generate_for_folder(folder, self.config)
            except Exception as e:
                print(f"Error starting document generation: {e}")

        messagebox.showinfo("Completed", f"Processing of folder {folder.name} completed.")

if __name__ == "__main__":
    ResolutionProcessor().run()
